import streamlit as st
import openpyxl
from openpyxl.styles import PatternFill
from PIL import Image
import io
import os

# --- Core Image to Excel Conversion Logic ---

def image_to_excel_pixel_art(image_data, should_resize, max_size, resampling_method):
    """
    Convert an image to Excel pixel art, with an option to skip resizing.
    
    Args:
        image_data (BytesIO or file-like object): Image data from user upload.
        should_resize (bool): If True, resize the image. Otherwise, use original size.
        max_size (int): The target size if resizing.
        resampling_method (Image.Resampling): The algorithm for resizing.
        
    Returns:
        BytesIO: An in-memory Excel file.
    """
    
    img = Image.open(image_data)
    
    if img.mode != 'RGB':
        img = img.convert('RGB')
    
    # Conditionally resize the image based on user's choice
    if should_resize:
        width, height = img.size
        if width > max_size or height > max_size:
            ratio = min(max_size / width, max_size / height)
            new_width = int(width * ratio)
            new_height = int(height * ratio)
            img = img.resize((new_width, new_height), resampling_method)
            st.info(f"Resized image from {width}x{height} to {new_width}x{new_height} for performance.")
    else:
        st.warning("Processing image at original size. This may be very slow.")

    width, height = img.size
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Pixel Art"
    
    for col_idx in range(1, width + 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 2.0
    
    for row_idx in range(1, height + 1):
        ws.row_dimensions[row_idx].height = 12.0
    
    fill_cache = {}
    
    progress_bar = st.progress(0, text="Processing pixels... (this can take a while for large images)")
    
    for y in range(height):
        for x in range(width):
            r, g, b = img.getpixel((x, y))
            hex_color = f"{r:02x}{g:02x}{b:02x}"
            
            if hex_color in fill_cache:
                fill = fill_cache[hex_color]
            else:
                fill = PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")
                fill_cache[hex_color] = fill
            
            cell = ws.cell(row=y + 1, column=x + 1)
            cell.fill = fill
        
        progress_percentage = (y + 1) / height
        progress_bar.progress(progress_percentage)
            
    ws.sheet_view.showGridLines = False
    
    progress_bar.progress(1.0, text="Finalizing Excel file...")

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    st.success("Conversion Complete!")
    
    return buffer

# --- Streamlit User Interface ---

st.set_page_config(page_title="Image to Excel Art", layout="centered")
st.title("🖼️ Image to Excel Pixel Art Converter")

# --- Sidebar for Settings ---
st.sidebar.header("⚙️ Settings")

# Add a toggle to enable/disable resizing
resize_image = st.sidebar.checkbox(
    "Resize image for performance", 
    value=True,
    help="Recommended. Uncheck to process the original image size, which can be VERY slow and may crash the app."
)

# Only show resolution/quality options if resizing is enabled
max_size = st.sidebar.slider(
    "Resolution (if resizing)", 
    min_value=32, 
    max_value=512,
    value=128,
    disabled=not resize_image,
    help="Higher values produce more detailed art but are slower."
)

resampling_options = {
    "Nearest (Blocky/Pixelated)": Image.Resampling.NEAREST,
    "Lanczos (Smoother/Detailed)": Image.Resampling.LANCZOS,
}
resampling_choice = st.sidebar.selectbox(
    "Resizing Quality (if resizing)",
    options=list(resampling_options.keys()),
    disabled=not resize_image,
    help="'Nearest' is blocky. 'Lanczos' is high-quality and smooth."
)
selected_method = resampling_options[resampling_choice]

# --- Main App Logic ---
uploaded_file = st.file_uploader(
    "Choose an image file",
    type=['png', 'jpg', 'jpeg', 'bmp', 'gif']
)

if uploaded_file is not None:
    # Display a strong warning if the user chooses not to resize a large image
    if not resize_image:
        # We need to peek at the image dimensions
        img_peek = Image.open(uploaded_file)
        width, height = img_peek.size
        # Rewind the file buffer so it can be read again later
        uploaded_file.seek(0)
        
        st.error(
            f"**WARNING:** You have chosen to process the original image size of **{width}x{height} pixels**. "
            f"This requires creating **{width * height:,}** individual cells in Excel. "
            "This process will be **EXTREMELY SLOW** and will likely **FAIL / TIME OUT** on the free cloud server. "
            "Proceed with caution."
        )

    st.image(uploaded_file, caption="Your Uploaded Image", use_column_width=True)
    
    if st.button("🎨 Convert to Excel Art"):
        with st.spinner("Processing... Please be patient, this might take several minutes."):
            try:
                # Pass the user's choices to the conversion function
                excel_buffer = image_to_excel_pixel_art(
                    uploaded_file, 
                    should_resize=resize_image, 
                    max_size=max_size, 
                    resampling_method=selected_method
                )
                
                original_filename = uploaded_file.name
                base_filename = os.path.splitext(original_filename)[0]
                size_str = f"{max_size}px" if resize_image else "original"
                excel_filename = f"{base_filename}_{size_str}_art.xlsx"

                st.download_button(
                    label="📥 Download Excel File",
                    data=excel_buffer,
                    file_name=excel_filename,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )

            except Exception as e:
                st.error(f"An error occurred: {e}")
                st.error("This can happen if the image is too large for the server's memory. Try enabling the 'Resize image' option in the sidebar.")
else:
    st.info("Upload an image to begin.")